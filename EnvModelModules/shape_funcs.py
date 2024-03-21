"""
#-------------------------------------------------------------------------------
# Name:
# Purpose:     Collection of functions relating to shapefiles
# Author:      Mike Martin
# Created:     11/12/2015
# Licence:     <your licence>
#-------------------------------------------------------------------------------
#
"""

__prog__ = 'shape_funcs.py'
__version__ = '0.0.0'
__author__ = 's03mm5'

from os.path import normpath, join, exists, isfile, isdir, basename
from os import listdir
from glob import glob
from shapefile import Reader
import math
from locale import setlocale, LC_ALL, format_string
from openpyxl import load_workbook
from netCDF4 import Dataset

ISO_LOOKUP = {'China': 'CHN', 'India': 'IND', 'United States': 'USA' }

class MakeBboxesNitroInpts(object, ):
    '''
    build bounding boxes for countries and provinces of large countries
    reconcile three naming conventions:
        GADM global administrative areas
        legacy inheritance from Astley-Shell
        FAO
    '''
    def __init__(self, settings, cntries_defn):
        '''

        '''
        # GADM bounding boxes by country - quite slow
        # ==============================
        cntry_bboxes = create_lookup_dict_gadm(settings)

        # read Excel file of N inputs
        # ===========================
        fao_glbl_n_inpts = _read_world_n_inputs(settings)
        major_states = list(fao_glbl_n_inpts.keys())  # India, China, the US
        major_states.remove('World')

        prvnc_bboxes = {}
        for cntry in major_states:

            # retrieve bounding boxes of provinces from GADM
            # ==============================================
            iso = ISO_LOOKUP[cntry]
            prvnc_bboxes[cntry] = _get_subdivisions(settings['shp_dir_gadm'], cntry, iso, 1)

        # Astley-Shell countries dataset
        # ==============================
        cntries_defn.nc_dset = Dataset(cntries_defn.nc_fname, mode='r')
        print('Opened dataset ' + cntries_defn.nc_fname)
        cntry_dict_shell = _shell_country_codes_lookup(fao_glbl_n_inpts)

        # harmonise GADM provinces with FAO
        # =================================
        prvnc_bboxes['China']['Xinjiang'] = prvnc_bboxes['China']['Xinjiang Uygur']
        prvnc_bboxes['China']['Ningxia']  = prvnc_bboxes['China']['Ningxia Hui']
        prvnc_bboxes['India']['Goa Daman & Diu'] = prvnc_bboxes['India']['Goa']

        prvnc_bboxes['India']['Delhi'] = prvnc_bboxes['India']['NCT of Delhi']
        prvnc_bboxes['India']['Jammu & Kashmir'] = prvnc_bboxes['India']['Jammu and Kashmir']
        prvnc_bboxes['India']['Pondicherry'] = prvnc_bboxes['India']['Puducherry']
        prvnc_bboxes['India']['Orissa'] = prvnc_bboxes['India']['Odisha']

        # populate object
        # ===============
        self.glbl_n_inpts = fao_glbl_n_inpts
        self.cntries_defn = cntries_defn
        self.major_states = major_states
        self.cntry_dict = cntry_dict_shell
        self.cntry_bboxes = cntry_bboxes
        self.prvnc_bboxes = prvnc_bboxes

        pass

def _read_world_n_inputs(settings):
    '''

    '''
    wb_obj = load_workbook(settings['n_inputs_xls'], data_only=True)
    global_n_inputs = {}

    for sht_name in wb_obj.sheetnames:
        sba_sht = wb_obj[sht_name]
        rows_generator = sba_sht.values
        for irow in range(3):
            dummy_row = next(rows_generator)
        data_rows = [list(row) for (_, row) in zip(range(400), rows_generator)]

        if sht_name  == 'US':
            cntry = 'United States'
        else:
            cntry = sht_name
        global_n_inputs[cntry] = data_rows

        if cntry == 'World':
            print('{} # countries:  {}'.format(cntry, len(data_rows)))
        else:
            print('Country: {}\t# provinces:  {}'.format(cntry, len(data_rows)))

    wb_obj.close()

    return global_n_inputs

def _get_subdivisions(shp_dir, country, iso_code, level, province = None, department = None, district = None,
                      canton = None):
    '''
    :param iso_code:
    :param level:
    :param province:
    :param department:
    :param district:
    :param canton:
    :return:

    get list of level 1 or 2 or 3 or 4 or 5 administrative names and corresponding bounding boxes
        level 1 consists of provinces/regions
        level 2 departments/prefectures
        level 3 counties/districts
        level 4 cantons
        level 5 communes
    '''
    admFname = 'gadm28_adm' + str(level) + '.shp'
    file_name = join(shp_dir, admFname)
    file_name = normpath(file_name)
    if not isfile(file_name):
        # return empty dictionary
        return {}

    # create Reader object
    # ====================
    shp_rdr = Reader(file_name)

    # collect record numbers for this country
    # =======================================
    if level == 1:
        print('Collecting record IDs for ' + country)
        rec_nums = []
        recs = shp_rdr.records()
        for recnum, record in enumerate(recs):
            if record[2] == iso_code:
                # obj_id = record[0]
                rec_nums.append(recnum)

        nrecs = len(rec_nums)
        if nrecs == 0:
            print('Country {} has no sub-divisions'.format(country))
        else:
            print('Found {} provinces/regions for {}'.format(nrecs,country))

        #
        subdiv_dict = {}
        for obj_id in rec_nums:
            province = recs[obj_id][5]
            bbox = shp_rdr.shape(obj_id).bbox
            area = calculate_area(bbox)
            bbox.append(area)
            subdiv_dict[province] = [round(float(val), 4) for val in bbox]

    # collect record numbers for this province
    # ========================================
    elif level == 2:
        print('Collecting record IDs for ' + province)
        rec_nums = []
        recs = shp_rdr.records()
        for recnum, record in enumerate(recs):
            if record[2] == iso_code and record[5] == province:
                # obj_id = record[0]
                rec_nums.append(recnum)

            # dbf file is in alphabetical order therefore we can finish the search when we hit next province
            if len(rec_nums) > 0 and record[5] != province:
                break

        nrecs = len(rec_nums)
        if nrecs == 0:
            print('Province {} has no sub-divisions'.format(province))
        else:
            print('Found {} departments/prefectures for {} province in {}'.format(nrecs, province, country))

        #
        subdiv_dict = {}
        for obj_id in rec_nums:
            department = recs[obj_id][7]
            bbox = shp_rdr.shape(obj_id).bbox
            area = calculate_area(bbox)
            bbox.append(area)
            subdiv_dict[department] = bbox

    return subdiv_dict

def create_lookup_dict_gadm(settings, show_message_flag = True, lenAdmNamesLim = 2):
    '''
    use data from GADM - global administrative areas (boundaries)
    see: https://gadm.org/
    '''
    shp_dir = settings['shp_dir_gadm']
    if not isdir(shp_dir):
        print('Shape directory ' + shp_dir + ' must exist')
        return {}
    max_countries = settings['max_countries']

    # there should be six GADM shapefiles (one for each level of administrative subdivision/aggregation)
    # ==================================================================================================
    admNames = glob(shp_dir + '/gadm28_adm*.shp')
    nadm_files = len(admNames)
    if nadm_files < 3:
        print('Insufficient GADM shapefiles in {} - should be at least three, found {}'.format(shp_dir, nadm_files))
        return {}

    # build dictionary of countries and corresponding three letter codes
    # ==================================================================
    cntry_lookup_dict = {}
    shp_rdr = Reader(admNames[0])
    print('Number of countries: {} - building country dictionary for maximum of {} countries...'
                                                                        .format(shp_rdr.numRecords, max_countries))
    for irec, shpe_rec in enumerate(shp_rdr.iterShapeRecords()):
        if irec > max_countries:
            break
        record = shpe_rec.record
        country_code, country = record[2:4]

        shpe = shpe_rec.shape
        bbox = [round(val, 4) for val in shpe.bbox]
        area = calculate_area(bbox)
        mess = 'Country: {}\tparts: {} points: {}'.format(country, len(shpe.parts), len(shpe.points))
        # print(mess)

        tmp_list = [country_code] + bbox + [round(area, 4), irec]
        cntry_lookup_dict.update({country:tmp_list})

    return cntry_lookup_dict

def create_lookup_dict(shp_dir, show_message_flag = True, lenAdmNamesLim = 2):

    # build dictionary of countries and corresponding three letter codes
    cntry_lookup_dict = {}
    if not isdir(shp_dir):
        print('Shape directory ' + shp_dir + ' must exist')
        return {}

    for country in listdir(shp_dir):   # get list of country folders in this directory e.g. China, Poland, UK
        country_path = join(shp_dir, country)
        if isdir(country_path):
            admNames = glob(country_path + '/*_adm*.shp')
            lenAdmNames = len(admNames)
            if lenAdmNames >= lenAdmNamesLim:

                country_code = basename(admNames[0])[0:3]

                # look for province level shape file then add to dictionary
                if show_message_flag:
                    print('Processing ' + country)
                shp_rdr = Reader(admNames[0])
                lon_ll, lat_ll, lon_ur, lat_ur = shp_rdr.bbox
                area = calculate_area(shp_rdr.bbox)

                tmp_list = [country_code] + [lon_ll, lat_ll, lon_ur, lat_ur, round(area), lenAdmNames]
                cntry_lookup_dict.update({country:tmp_list})

    return cntry_lookup_dict

def _getPolys(lgr,shpes):

    func_name =  __prog__ + ' _getPolys'


    lgr.info('Function: {0}\tnumber of shapes: {1}'.format(func_name,len(shpes)))
    # for now just return the first polygon

    return shpes[-1].points

def _getPoly(lgr,shpe):

    func_name =  __prog__ + ' _getPoly'
    # stand in - should return bounding boxes
    parts = shpe.parts

    lgr.info('Function: {0}\tnumber of parts: {1}'.format(func_name,len(parts)))
    # for now just return the first polygon
    if len(parts) > 1:
        return shpe.points[0: parts[1]]
    else:
        return shpe.points

def get_shape_file_info(shp_file):

    func_name =  __prog__ + '  get_shape_file_info'

    if not isfile(shp_file):
        return [], []

    # create Reader object and pass the name of an existing shapefile
    shp_rdr = Reader(shp_file)
    # bbox = shp_rdr.bbox
    # flds = shp_rdr.fields
    shpes = shp_rdr.shapes()

    recs = shp_rdr.records()
    nrecs = len(recs); nshpes = len(shpes)
    if nshpes != nrecs:
        print('Number of shapes {0:d} does not match number of records {1:d} in DBF file for shapefile {2:s}'
              .format(nshpes,nrecs,shp_file))
        return [], []

    return shpes, recs

def _getShpe(shp_dir, country, abbrev, level, province = None, department = None,
                                        district = None, canton = None):
    """
    # get required polygon for a given province or department
    # level 1 consists of provinces/regions
    # level 2 departments/prefectures
    # level 3 counties/districts
    # level 4 cantons
    # level 5 communes
    """
    func_name =  __prog__ + ' _getShpe'

    admFname = '_adm' + str(level) + '.shp'
    file_name = join(shp_dir + '/' + country, abbrev + admFname)
    if not isfile(file_name):
        return -1

    # create Reader object and pass the name of an existing shapefile
    shp_rdr = Reader(file_name)
    # bbox = shp_rdr.bbox
    # flds = shp_rdr.fields
    shpes = shp_rdr.shapes()

    recs = shp_rdr.records()
    nrecs = len(recs); nshpes = len(shpes)
    if nshpes != nrecs:
        print('Number of shapes {0:d} does not match number of records {1:d} in DBF file for shapefile {2:s}'
              .format(nshpes,nrecs,file_name))
        return -1
    # construct dictionary consisting of names of areas (e.g. province/department/district etc.)
    # and corresponding bounding boxes
    subdiv_dict = {}
    # extract names
    shpe = -1
    for irec in range(nrecs):
        subdiv = None  # default
        # country level - the coarsest level therefore take first record
        if level == 0:
            shpe = shpes[irec]
            break
        # province level - the next coarsest level therefore no requirement to check other fields
        elif level == 1:
            if province == recs[irec][4]:
                # poly = shpes[irec].points
                # parts = shpes[irec].parts
                shpe = shpes[irec]
                break
        # department/prefecture - therefore filter out irrelevant i.e. non-parent provinces
        elif level == 2:
            if province == recs[irec][4] and department == recs[irec][6]:
                # poly = shpes[irec].points
                # parts = shpes[irec].parts
                shpe = shpes[irec]
                break
        elif level == 3:
            if province == recs[irec][4] and department == recs[irec][6] and district == recs[irec][8]:
                # poly = shpes[irec].points
                # parts = shpes[irec].parts
                shpe = shpes[irec]
                break  # field name NAME_3, column I
        elif level == 4:
            if province == recs[irec][4] and department == recs[irec][6] and \
                                district == recs[irec][8] and canton == recs[irec][10]:
                # poly = shpes[irec].points
                # parts = shpes[irec].parts
                shpe = shpes[irec]
                break
        elif level == 5:
                pass
    if shpe == -1:
        print('Func: {0} problem: Could not find province: {1}\tdepartment: {2}\tdistrict: {3}\tlevel: {4}'.
                                format(func_name, province, department, district, level))
    else:
        print('Leaving {0}\tlevel: {1}, number of parts: {2}\tpoints on polygon: {3}\trecord: {4}'.
                                format(func_name, level, len(shpe.parts), len(shpe.points), irec))
    return shpe


def get_subdivisions(shp_dir, country, iso_code, level, province = None, department = None, district = None,
                                                                                                canton = None):
    '''
    get list of level 1 or 2 or 3 or 4 or 5 administrative names and corresponding bounding boxes
        level 1 consists of provinces/regions
        level 2 departments/prefectures
        level 3 counties/districts
        level 4 cantons
        level 5 communes
    '''
    admFname = 'gadm28_adm' + str(level) + '.shp'
    file_name = join(shp_dir, admFname)
    file_name = normpath(file_name)
    if not isfile(file_name):
        # return empty dictionary
        return {}

    # create Reader object
    # ====================
    shp_rdr = Reader(file_name)

    # collect record numbers for this country
    # =======================================
    if level == 1:
        print('Collecting record IDs for ' + country)
        rec_nums = []
        recs = shp_rdr.records()
        for recnum, record in enumerate(recs):
            if record[2] == iso_code:
                # obj_id = record[0]
                rec_nums.append(recnum)

        nrecs = len(rec_nums)
        if nrecs == 0:
            print('Country {} has no sub-divisions'.format(country))
        else:
            print('Found {} provinces/regions for {}'.format(nrecs,country))

        #
        subdiv_dict = {}
        for obj_id in rec_nums:
            province = recs[obj_id][5]
            bbox = shp_rdr.shape(obj_id).bbox
            area = calculate_area(bbox)
            bbox.append(area)
            subdiv_dict[province] = bbox

    # collect record numbers for this province
    # ========================================
    elif level == 2:
        print('Collecting record IDs for ' + province)
        rec_nums = []
        recs = shp_rdr.records()
        for recnum, record in enumerate(recs):
            if record[2] == iso_code and record[5] == province:
                # obj_id = record[0]
                rec_nums.append(recnum)

            # dbf file is in alphabetical order therefore we can finish the search when we hit next province
            if len(rec_nums) > 0 and record[5] != province:
                break

        nrecs = len(rec_nums)
        if nrecs == 0:
            print('Province {} has no sub-divisions'.format(province))
        else:
            print('Found {} departments/prefectures for {} province in {}'.format(nrecs, province, country))

        #
        subdiv_dict = {}
        for obj_id in rec_nums:
            department = recs[obj_id][7]
            bbox = shp_rdr.shape(obj_id).bbox
            area = calculate_area(bbox)
            bbox.append(area)
            subdiv_dict[department] = bbox

    return subdiv_dict

def get_shps_provinces(shp_file):

    if not exists(shp_file):
        return ''

     # create Reader object and pass the name of an existing shapefile
    shp_rdr = Reader(shp_file)
    # bbox = shp_rdr.bbox
    # flds = shp_rdr.fields
    shpes = shp_rdr.shapes()

    recs = shp_rdr.records()
    nrecs = len(recs); nshpes = len(shpes)
    if nshpes != nrecs:
        print('Number of shapes {0:d} does not match number of records {1:d} in DBF file for shapefile {2:s}'
              .format(nshpes,nrecs,shp_file))
        return -1

    provinces = []
    for rec in recs:
        prov = rec[1]
        if prov not in provinces:
            provinces.append(prov)

    nprovs = len(provinces)

    return '{}, {}'.format(nshpes, nprovs)

def calculate_area(bbox):
    '''
    calculate Zone of earth's sphere i.e. the curved surface of a spherical segment.
    '''
    radius_earth = 6371.0 # km average
    pi = 3.14159265
    deg2rad = pi/180.0
    lon_ll, lat_ll, lon_ur, lat_ur = bbox

    # TODO: work out what can happen for southern hemisphere or straddling equator....
    theta2 = lat_ur*deg2rad
    theta1 = lat_ll*deg2rad

    domega = (lon_ur - lon_ll)/360.0

    # area is in square km
    # ====================
    area = 2.0*pi*domega*radius_earth*(radius_earth*(math.sin(theta2) - math.sin(theta1)))

    return area

def format_bbox(bbox, area = 0.0, ndigits = 0):
    """
    format long/lat and area string for display in GUI
    """

    if len(bbox) == 7:
        abbrev, lon_ll, lat_ll, lon_ur, lat_ur, area, nlevels = bbox
    elif len(bbox) == 5:
        lon_ll, lat_ll, lon_ur, lat_ur, area = bbox
    else:
        lon_ll, lat_ll, lon_ur, lat_ur = bbox

    if ndigits == 0:
        setlocale(LC_ALL, '')
        area_out = round(area)
        area_out = format_string("%d", area_out, grouping=True )
    else:
        area_out = round(area, ndigits)

    return 'LL: {:.2f} {:.2f}\tUR: {:.2f} {:.2f}\tArea: {} km\u00b2'.format(lon_ll, lat_ll, lon_ur, lat_ur, area_out)

def compress_shpe_file(shp_file, images_dir):

    func_name =  __prog__ + ' compress_shpe_file'

    if not isfile(shp_file):
        return -1

    # create Reader object and pass the name of an existing shapefile
    shp_rdr = Reader(shp_file)
    # bbox = shp_rdr.bbox
    # flds = shp_rdr.fields
    shpes = shp_rdr.shapes()

    recs = shp_rdr.records()
    nrecs = len(recs); nshpes = len(shpes)
    if nshpes != nrecs:
        print('Number of shapes {0:d} does not match number of records {1:d} in DBF file for shapefile {2:s}'
              .format(nshpes,nrecs,shp_file))
        return -1

    # file to output first ring
    fname = join(images_dir,'border.csv')
    if not isfile(fname):
        print('Function: {0}\tfile: {1} does not exist - will create'.format(func_name,fname))
    try:
        fpoly = open(fname,'w')
    except PermissionError:
        print('Function: {0}\tCould not open file: {1}'.format(func_name,fname))
        return 1

    # fetch each shape and map to granular
    granularity = 120.0
    coord_max_val = 999.0
    for nshp, shpe in zip(range(len(shpes)), shpes):
        if nshp > 0:
            break
        print('shape {} with {} parts and {} points'.format(nshp + 1, len(shpe.parts), len(shpe.points)))

        indxs = 0
        for npart, indxe in zip(range(len(shpe.parts)),shpe.parts[1:]):
            if npart > 0:
                break
            lat_max = -coord_max_val
            lat_min = coord_max_val
            lon_max = -coord_max_val
            lon_min = coord_max_val
            gran_points = []
            ring = shpe.points[indxs:indxe ]
            print('ring {} first/last points: {} {}'.format(npart + 1, ring[0],ring[-1]))
            for point in ring:
                lon, lat = point
                if lat > lat_max:
                    lat_max = lat
                if lat < lat_min:
                    lat_min = lat
                if lon > lon_max:
                    lon_max = lon
                if lon < lon_min:
                    lon_min = lon
                nrow = round((90.0 -  lat)*granularity)
                ncol = round((180.0 + lon)*granularity)
                elem = list([ncol,nrow])
                if elem not in gran_points:
                    gran_points.append(elem)
                    fpoly.write('{0},{1},{2},{3}\n'.format(lon,lat,nrow,ncol,nshp))

            print('ring {} with {} points, granular points: {} min/max lon: {} {} min/max lat: {} {}'
                  .format(npart + 1,indxe - indxs, len(gran_points), lon_min, lon_max, lat_min, lat_max))
            indxs = indxe
    fpoly.close()

    print('Leaving {0}'.format(func_name))
    return shpe

def _shell_country_codes_lookup(glbl_n_inpts):
    '''
    from miscanfor.f90
    '''
    acountries = [58, 32, 11, 218, 165, 183, 232, 107, 54, 173, 200, 22, 101, 75, 72, 133, 70, 39, 18, 112, 17, 64, 1,
                  31, 33, 143, 66, 38, 13, 42, 160, 151, 241, 242, 26, 25, 87, 216, 117, 188, 153, 152, 139, 12, 123,
                  115, 71, 23, 211, 53, 68, 78, 222, 138, 167, 113, 150, 240, 111, 239, 163, 210, 213, 88, 230, 5, 149,
                  7, 52, 110, 14, 99, 77, 102, 170, 0, 145, 144, 142, 76, 114, 27, 80, 100, 221, 129, 51, 79, 225, 47,
                  234, 84, 83, 192, 48, 55, 157, 82, 136, 132, 61, 65, 36, 49, 28, 10, 21, 219, 35, 19, 215, 34, 105,
                  46, 98, 9, 122, 171, 120, 187, 103, 41, 217, 63, 8, 154, 134, 164, 62, 174, 176, 229, 202, 116, 16,
                  147, 109, 130, 59, 166, 140, 137, 2, 141, 89, 108, 104, 168, 74, 124, 97, 43, 69, 207, 172, 20, 148,
                  127, 119, 146, 73, 29, 24, 94, 156, 106, 81, 85, 126, 4, 45, 86, 220, 50, 92, 60, 155, 93, 15, 91,
                  121, 90, 56, 205, 44, 125, 131, 226, 228, 3, 57, 161, 175, 135, 236, 6, 128, 30, 37, 40]

    # where necessary these names have been changed to names in N inputs Excel file
    # =============================================================================
    cnames = ['Afghanistan', 'Albania', 'Algeria', 'Spain', 'Angola', 'Anguilla', 'Antigua and Barbuda',
              'Argentina'] + \
             ['Armenia', 'Venezuela', 'Australia', 'Austria', 'Russian Federation', 'Bahamas', 'Bahrain', 'Bangladesh',
              'Barbados', 'Belarus', 'Belgium', 'Belize'] + \
             ['Benin', 'Bhutan', 'Bolivia',  'Bosnia and Herzegovina', 'Botswana', 'Brazil', 'Brunei Darussalam',
              'Bulgaria', 'Burkina Faso', 'Burundi', 'Cambodia', 'Cameroon'] + \
             ['Canada', 'Canada',  'Central African Republic', 'Chad', 'Chile', 'China', 'Colombia', 'Comoros',
              'Dem Republic of Congo', 'Republic of Congo'] + \
             ['Costa Rica', 'Cote d\'Ivoire', 'Croatia', 'Cuba', 'Cyprus', 'Czech Republic', 'Denmark', 'Djibouti',
              'Dominica', 'Dominican Republic', 'East Timor', 'Ecuador'] + \
             ['Egypt', 'El Salvador', 'Equatorial Guinea', 'Eritrea', 'Estonia', 'Ethiopia', 'Argentina',
              'Faroe Islands', 'New Caledonia', 'Finland', 'France', 'French Guiana'] + \
             ['Gabon', 'Gambia', 'Georgia', 'Germany', 'Ghana', 'Greece', 'Greenland', 'Grenada', 'Guadeloupe',
              'Guatemala', 'Guinea', 'Guinea-Bissau', 'Guyana', 'Haiti', 'Honduras'] + \
             ['Hungary', 'Iceland', 'India', 'Indonesia', 'Iran Islamic Republic', 'Iraq', 'Ireland', 'United Kingdom',
              'Israel incl. Palestine', 'Italy',
              'Jamaica', 'Japan', 'United Kingdom', 'Jordan', 'Kazakhstan', 'Kenya'] + \
             ['Dem People\'s Rep Korea', 'Republic of Korea', 'Kuwait', 'Kyrgyzstan', 'Laos', 'Latvia', 'Lebanon', 'Lesotho',
              'Liberia', 'Libyan Arab Jamahiriya', 'Switzerland', 'Lithuania', 'Luxembourg'] + \
             ['Macau', 'Macedonia', 'Madagascar', 'Malawi', 'Malaysia', 'Mali', 'Malta', 'Martinique', 'Mauritania',
              'Mayotte', 'Mexico', 'Russian Federation', 'France', 'Mongolia', 'Morocco'] + \
             ['Mozambique', 'Myanmar', 'Namibia', 'Nepal', 'Netherlands', 'Netherlands Antilles', 'New Caledonia',
              'New Zealand', 'Nicaragua', 'Niger', 'Nigeria', 'Norway', 'Oman'] + \
             ['Pakistan', 'Israel incl. Palestine', 'Panama', 'Papua New Guinea', 'Paraguay', 'Peru', 'Philippines', 'Poland',
              'Portugal', 'Puerto Rico', 'Kuwait', 'Romania', 'Russian Federation', 'Rwanda'] + \
             ['Saint Lucia', 'Saint Pierre and Miquelon', 'Saint Vincent and the Grenadines', 'San Marino',
              'Sao Tome and Principe', 'Saudi Arabia', 'Senegal'] + \
             ['Sierra Leone', 'Singapore', 'Slovak Republic', 'Slovenia', 'Soloman Islands', 'Somalia', 'South Africa',
              'Spain', 'Sri Lanka', 'Sudan', 'Suriname', 'Swaziland'] + \
             ['Sweden', 'Switzerland', 'Syrian Arab Republic', 'China', 'Tajikistan', 'United Rep of Tanzania',
              'Thailand', 'Togo',
              'Trinidad and Tobago', 'Tunisia', 'Turkey', 'Turkmenistan'] + \
             ['Turks and Caicos Islands', 'Uganda', 'Ukraine', 'United Arab Emirates', 'United Kingdom',
              'United States', 'Uruguay', 'Uzbekistan', 'Vanuatu'] + \
             ['Venezuela', 'Vietnam', 'Virgin Islands (U.S.)', 'Morocco', 'Yemen', 'Serbia and Montenegro', 'Zambia',
              'Zimbabwe']

    # collect country names
    # =====================
    n_inpt_cntrys = []
    for rec in glbl_n_inpts['World']:
        cntry = rec[0]
        if cntry == 'WORLD':
            continue
        n_inpt_cntrys.append(cntry)
    n_inpt_cntrys = sorted(n_inpt_cntrys)

    print('Found ' + str(len(n_inpt_cntrys)) + ' countries in N inputs Excel file')

    # map Astley-Shell names
    # ======================
    cntry_dict = {}
    not_founds = []
    for icntry, cname in zip(acountries, cnames):
        if cname in n_inpt_cntrys:
            cntry_dict[icntry] = cname
        else:
            not_founds.append(cname)

    n_not_fnds = len(not_founds)
    if n_not_fnds > 0:
        print(str(n_not_fnds) + ' countries from the Astley-Shell names file not found in N inputs Excel file:')
        print('\t' + str(not_founds))

    return cntry_dict

